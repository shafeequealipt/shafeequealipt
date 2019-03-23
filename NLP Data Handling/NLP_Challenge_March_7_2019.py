import re
import time
from collections import Counter
from itertools import groupby

import nltk
import nltk.tag.stanford as st
from nltk.corpus import stopwords
from nltk.tokenize import RegexpTokenizer
from nltk.tokenize import word_tokenize
from openpyxl import Workbook
from openpyxl import load_workbook


def remove_work_notes(data_string):
    invalid_string = 'work notes'
    data_string_lines = data_string.splitlines()
    cleaned_string_lines = [dataLine for dataLine in data_string_lines if
                            invalid_string.lower() not in dataLine.lower()]
    return '\n'.join(cleaned_string_lines)


def remove_hyperlinks(data_string):
    return re.sub(r'\w+:\/{2}[\d\w-]+(\.[\d\w-]+)*(?:(?:\/[^\s/]*))*', '', data_string)


def remove_stop_words(data_string):
    stop_words = set(stopwords.words('english'))
    word_tokens = tokenizer.tokenize(data_string.lower())
    filtered_sentence = [w for w in word_tokens if w not in stop_words]
    return filtered_sentence


def fetch_kb_references(data_string):
    regex = r'KB\d{4,}'
    matches = re.findall(regex, data_string, re.MULTILINE)
    return matches, len(matches) > 0


def find_person_names(data_string):
    word_tokens = word_tokenize(data_string)
    classified_text = tagger.tag(word_tokens)
    person_list = []
    for tag, chunk in groupby(classified_text, lambda x: x[1]):
        if tag == "PERSON":
            person_list.append(" ".join(w for w, t in chunk))
    return person_list


def generate_grams(data_string):
    try:
        tokens = tokenizer.tokenize(data_string.lower())
        bigram = nltk.bigrams(tokens)
        trigram = nltk.trigrams(tokens)
        return find_top10_words(data_string), Counter(bigram).most_common(10), Counter(trigram).most_common(10)
    except Exception as e:
        return '', '', ''


def remove_special_characters(data_string):
    regex = r'[^a-zA-Z0-9\s]{4,}'
    data_string = re.sub(regex, ' ', data_string)
    regex = r'^[-=]'
    return re.sub(regex, '', data_string)


# In[10]:


def find_top10_words(data_string):
    words = re.findall(r'\w+', data_string.lower())
    return Counter(words).most_common(10)


def process_data(input_data_string):
    output_data = remove_work_notes(input_data_string)
    output_data = remove_hyperlinks(output_data)
    output_data = remove_special_characters(output_data)
    stop_words_removed_string = ' '.join(remove_stop_words(output_data))
    # preparing output parameters
    kb_references, has_kb_reference = fetch_kb_references(input_data_string)
    kb_info = 'No'
    if has_kb_reference:
        kb_info = 'Yes ' + str(kb_references)
    person_names = find_person_names(input_data_string)
    uni, bi, tri = generate_grams(stop_words_removed_string)
    return {
        'Data': output_data,
        'KB Reference': kb_info,
        'Person Names': ', '.join(set(person_names)),
        'Unigram Frequency': uni,
        "Bigram Frequency": bi,
        "Trigram Frequency": tri
    }


print('Processing file.')
# setting parameters
inputFilePath ='NLP_Challenge_07_Mar.xlsx'
inputSheetName = 'Sheet1'
inputDataColumn = 'A'
tableHasHeader = True
outputFilePath = 'NLP_Challenge_07_Mar_Output.xlsx'
outputSheetName = 'Sheet1'
tokenizer = RegexpTokenizer(r'\w+')
# ner settings
gzPath = 'stanford-ner\\english.all.3class.distsim.crf.ser.gz'
jarPath = 'stanford-ner\\stanford-ner.jar'
tagger = st.StanfordNERTagger(gzPath, jarPath)

# read input data
print('Reading input data.')
inputWorkbook = load_workbook(filename=inputFilePath, read_only=False)
inputWorksheet = inputWorkbook[inputSheetName]
inputWorksheetRowCount = inputWorksheet.max_row
inputData = [col.value for col in inputWorksheet['A']]
if tableHasHeader and len(inputData) > 0:
    inputData.pop(0)
print('Processing input data.')
processedData = [process_data(data) for data in inputData]

print('Writing processed data to output excel file.')
outputWorkbook = Workbook()
outputSheet = outputWorkbook.create_sheet('Output Data', 0)
outputSheet.append(list(processedData[0].keys()))

for data in processedData:
    outputSheet.append([str(data[k]) for k in data])

time_string = time.strftime("%Y%m%d_%H%M%S")
outputWorkbook.save(f'output_at_{time_string}.xlsx')
print('Process completed.')
